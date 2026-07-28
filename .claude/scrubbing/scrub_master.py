#!/usr/bin/env python3
"""
scrub_master.py — the one writer for every LTP scrub master workbook.

Why this exists: the scrubbers were creating a new sheet on every run, so the
current list was impossible to find. Sheet naming and append-vs-create are now
decided by this script, not by a model remembering a rule.

Guarantees, on every run:
  * One master workbook per vertical. Never a second file.
  * One sheet per STATE, named with the 2-letter code (NY, TX, NC). Never a
    city name, never "NY (2)", never a date-stamped tab.
  * A run against a state that already has a sheet APPENDS to that sheet.
  * Duplicate rows are dropped on (business name + city + state).
  * The master is copied into _versions/ BEFORE it is touched, so any run can
    be reverted.

Usage
-----
  # add rows (rows come in as JSON: a list of objects keyed by column header)
  python3 scrub_master.py add --vertical collision --state NY --rows rows.json

  # see what's in a master
  python3 scrub_master.py status --vertical collision

  # list restore points
  python3 scrub_master.py versions --vertical collision

  # undo the last run (or a specific version)
  python3 scrub_master.py revert --vertical collision
  python3 scrub_master.py revert --vertical collision --version "2026-07-28 141503"

  # one-time: fold existing city/junk sheets into proper state sheets
  python3 scrub_master.py consolidate --vertical collision

Set the folder once, either by editing CLAUDE_FOLDER below or by exporting
SCRUBBING_ROOT.
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")

# --- where things live -------------------------------------------------------
# The "Claude" folder in Finder. Override with:  export SCRUBBING_ROOT=...
CLAUDE_FOLDER = Path(
    os.environ.get("SCRUBBING_ROOT", Path.home() / "Claude" / "Scrubbing Docs")
).expanduser()

VERSIONS_TO_KEEP = 20

STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY","DC",
}

STATE_NAMES = {
    "alabama":"AL","alaska":"AK","arizona":"AZ","arkansas":"AR","california":"CA",
    "colorado":"CO","connecticut":"CT","delaware":"DE","florida":"FL","georgia":"GA",
    "hawaii":"HI","idaho":"ID","illinois":"IL","indiana":"IN","iowa":"IA",
    "kansas":"KS","kentucky":"KY","louisiana":"LA","maine":"ME","maryland":"MD",
    "massachusetts":"MA","michigan":"MI","minnesota":"MN","mississippi":"MS",
    "missouri":"MO","montana":"MT","nebraska":"NE","nevada":"NV",
    "new hampshire":"NH","new jersey":"NJ","new mexico":"NM","new york":"NY",
    "north carolina":"NC","north dakota":"ND","ohio":"OH","oklahoma":"OK",
    "oregon":"OR","pennsylvania":"PA","rhode island":"RI","south carolina":"SC",
    "south dakota":"SD","tennessee":"TN","texas":"TX","utah":"UT","vermont":"VT",
    "virginia":"VA","washington":"WA","west virginia":"WV","wisconsin":"WI",
    "wyoming":"WY","district of columbia":"DC",
}

# --- the four verticals ------------------------------------------------------
# "key" columns form the dedupe identity. Keep these stable.
VERTICALS = {
    "childcare": {
        "workbook": "Childcare Master.xlsx",
        "columns": [
            "Company Name", "Street", "City", "State", "Licensed Capacity",
            "Owner First Name", "Owner Last Name", "Notes",
        ],
        "key": ["Company Name", "City", "State"],
    },
    "medspa": {
        "workbook": "MedSpa Master.xlsx",
        "columns": [
            "Business Name", "Units", "Salutation", "First name", "Last name",
            "Street address", "Suite / unit", "City", "State", "Zip code",
            "Mailer sent", "Date", "Notes",
        ],
        "key": ["Business Name", "City", "State"],
    },
    "collision": {
        "workbook": "Collision Master.xlsx",
        "columns": [
            "Business Name", "Units", "Salutation", "First name", "Last name",
            "Street address", "Suite / unit", "City", "State", "Zip code",
            "Mailer sent", "Date", "Notes",
        ],
        "key": ["Business Name", "City", "State"],
    },
    "oil": {
        "workbook": "Oil Change Master.xlsx",
        "columns": [
            "Business Name", "Units", "Salutation", "First name", "Last name",
            "Street address", "Suite / unit", "City", "State", "Zip code",
            "Mailer sent", "Date", "Notes",
        ],
        "key": ["Business Name", "City", "State"],
    },
}

ALIASES = {
    "child care": "childcare", "daycare": "childcare", "preschool": "childcare",
    "med spa": "medspa", "medical spa": "medspa", "med-spa": "medspa",
    "collision repair": "collision", "body shop": "collision", "auto body": "collision",
    "oil change": "oil", "quick lube": "oil", "lube": "oil", "sboc": "oil",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


# --- helpers -----------------------------------------------------------------
def resolve_vertical(name: str) -> str:
    n = (name or "").strip().lower()
    n = ALIASES.get(n, n)
    if n not in VERTICALS:
        sys.exit(f"Unknown vertical {name!r}. Pick one of: {', '.join(VERTICALS)}")
    return n


def resolve_state(raw: str) -> str:
    """Accept 'NY', 'ny', 'New York', or 'Buffalo, NY' — always return 'NY'.

    A city is never a sheet name. If a city is passed, its state is what counts.
    """
    s = (raw or "").strip()
    if not s:
        sys.exit("A state is required. Sheets are named by state, never by city.")
    if "," in s:                       # "Buffalo, NY"
        s = s.rsplit(",", 1)[-1].strip()
    up = s.upper()
    if up in STATES:
        return up
    low = s.lower()
    if low in STATE_NAMES:
        return STATE_NAMES[low]
    sys.exit(
        f"{raw!r} is not a US state. Sheets are named by 2-letter state code.\n"
        f"If you ran a city, pass the state that city is in."
    )


def norm(v) -> str:
    """Normalize a cell for dedupe: casefold, strip punctuation and legal suffixes."""
    s = str(v or "").strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b(inc|llc|ltd|co|corp|company|the)\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def paths(vertical: str):
    spec = VERTICALS[vertical]
    folder = CLAUDE_FOLDER / vertical.capitalize() if False else CLAUDE_FOLDER / {
        "childcare": "Childcare", "medspa": "MedSpa",
        "collision": "Collision", "oil": "Oil",
    }[vertical]
    return folder, folder / spec["workbook"], folder / "_versions"


def style_sheet(ws, columns):
    for i, name in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        width = 42 if name == "Notes" else max(12, min(len(name) + 8, 30))
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def open_master(vertical: str):
    folder, master, versions = paths(vertical)
    folder.mkdir(parents=True, exist_ok=True)
    versions.mkdir(parents=True, exist_ok=True)
    if master.exists():
        return load_workbook(master)
    wb = Workbook()
    wb.remove(wb.active)          # no stray "Sheet"
    return wb


def snapshot(vertical: str, label: str) -> Path | None:
    """Copy the master into _versions/ before modifying it."""
    _, master, versions = paths(vertical)
    if not master.exists():
        return None
    versions.mkdir(parents=True, exist_ok=True)   # consolidate/revert may run first
    stamp = datetime.now().strftime("%Y-%m-%d %H%M%S")
    dest = versions / f"{master.stem} — {stamp} ({label}).xlsx"
    shutil.copy2(master, dest)
    old = sorted(versions.glob(f"{master.stem} — *.xlsx"))
    for stale in old[:-VERSIONS_TO_KEEP]:
        stale.unlink(missing_ok=True)
    return dest


def sheet_for_state(wb, state: str, columns):
    """Return the sheet for this state, creating it ONLY if absent.

    This is the whole point of the script. Never make a second sheet for a
    state that already has one.
    """
    for ws in wb.worksheets:
        if ws.title.strip().upper() == state:
            return ws, False
    ws = wb.create_sheet(title=state)
    style_sheet(ws, columns)
    return ws, True


def existing_keys(ws, columns, key_cols):
    headers = [c.value for c in ws[1]]
    idx = [headers.index(k) for k in key_cols if k in headers]
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(row):
            seen.add(tuple(norm(row[i]) if i < len(row) else "" for i in idx))
    return seen


def order_sheets(wb):
    wb._sheets.sort(key=lambda ws: ws.title.upper())


# --- commands ----------------------------------------------------------------
def cmd_add(args):
    vertical = resolve_vertical(args.vertical)
    state = resolve_state(args.state)
    spec = VERTICALS[vertical]
    columns, key_cols = spec["columns"], spec["key"]

    raw = json.loads(Path(args.rows).read_text()) if args.rows != "-" else json.load(sys.stdin)
    if isinstance(raw, dict):
        raw = raw.get("rows", [])
    if not isinstance(raw, list):
        sys.exit("rows must be a JSON list of objects keyed by column header.")

    folder, master, _ = paths(vertical)
    snap = snapshot(vertical, f"before {state}")
    wb = open_master(vertical)
    ws, created = sheet_for_state(wb, state, columns)
    seen = existing_keys(ws, columns, key_cols)

    added = skipped = 0
    for r in raw:
        if not isinstance(r, dict):
            continue
        r = {k: v for k, v in r.items()}
        r["State"] = state                      # the sheet is the source of truth
        k = tuple(norm(r.get(c, "")) for c in key_cols)
        if k in seen:
            skipped += 1
            continue
        seen.add(k)
        ws.append([r.get(c, "") for c in columns])
        added += 1

    order_sheets(wb)
    wb.save(master)

    verb = "created" if created else "appended to"
    print(f"{vertical}: {verb} sheet '{state}' — {added} added, {skipped} duplicate(s) skipped")
    print(f"  master:  {master}")
    print(f"  sheets:  {', '.join(w.title for w in wb.worksheets)}")
    if snap:
        print(f"  revert:  python3 scrub_master.py revert --vertical {vertical}")


def cmd_status(args):
    vertical = resolve_vertical(args.vertical)
    folder, master, _ = paths(vertical)
    if not master.exists():
        print(f"{vertical}: no master yet at {master}")
        return
    wb = load_workbook(master, read_only=True)
    print(f"{vertical}: {master}")
    total = 0
    for ws in wb.worksheets:
        n = max(ws.max_row - 1, 0)
        total += n
        flag = "" if ws.title.upper() in STATES else "   <-- NOT a state sheet, run consolidate"
        print(f"  {ws.title:<6} {n:>6} rows{flag}")
    print(f"  {'TOTAL':<6} {total:>6} rows")


def cmd_versions(args):
    vertical = resolve_vertical(args.vertical)
    _, master, versions = paths(vertical)
    found = sorted(versions.glob(f"{master.stem} — *.xlsx"), reverse=True)
    if not found:
        print(f"{vertical}: no restore points yet")
        return
    print(f"{vertical}: {len(found)} restore point(s), newest first")
    for p in found:
        print(f"  {p.name}")


def cmd_revert(args):
    vertical = resolve_vertical(args.vertical)
    _, master, versions = paths(vertical)
    found = sorted(versions.glob(f"{master.stem} — *.xlsx"), reverse=True)
    if not found:
        sys.exit(f"{vertical}: nothing to revert to.")
    if args.version:
        match = [p for p in found if args.version in p.name]
        if not match:
            sys.exit(f"No restore point matching {args.version!r}. Try: versions")
        pick = match[0]
    else:
        pick = found[0]
    snapshot(vertical, "before revert")          # reverting is itself undoable
    shutil.copy2(pick, master)
    print(f"{vertical}: reverted to {pick.name}")
    print(f"  master: {master}")


def cmd_consolidate(args):
    """Fold city-named / stray sheets into the right state sheet, once."""
    vertical = resolve_vertical(args.vertical)
    spec = VERTICALS[vertical]
    columns, key_cols = spec["columns"], spec["key"]
    folder, master, _ = paths(vertical)
    if not master.exists():
        sys.exit(f"{vertical}: no master at {master}")

    snapshot(vertical, "before consolidate")
    wb = load_workbook(master)
    moved, dropped, kept = 0, 0, []

    for ws in list(wb.worksheets):
        title = ws.title.strip()
        if title.upper() in STATES:
            kept.append(title.upper())
            if ws.title != title.upper():
                ws.title = title.upper()
            continue
        headers = [c.value for c in ws[1]]
        if "State" not in headers:
            print(f"  ? '{title}' has no State column — leaving it alone")
            continue
        si = headers.index("State")
        rows_by_state = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            st = str(row[si] or "").strip().upper()
            if st not in STATES:
                continue
            rows_by_state.setdefault(st, []).append(
                {h: row[i] if i < len(row) else "" for i, h in enumerate(headers)}
            )
        for st, rows in rows_by_state.items():
            target, _ = sheet_for_state(wb, st, columns)
            seen = existing_keys(target, columns, key_cols)
            for r in rows:
                k = tuple(norm(r.get(c, "")) for c in key_cols)
                if k in seen:
                    dropped += 1
                    continue
                seen.add(k)
                target.append([r.get(c, "") for c in columns])
                moved += 1
        wb.remove(ws)
        print(f"  merged '{title}' -> {', '.join(sorted(rows_by_state)) or '(no usable rows)'}")

    order_sheets(wb)
    wb.save(master)
    print(f"{vertical}: {moved} row(s) merged, {dropped} duplicate(s) dropped")
    print(f"  sheets now: {', '.join(w.title for w in wb.worksheets)}")


def main():
    p = argparse.ArgumentParser(description="One master workbook per vertical, one sheet per state.")
    sub = p.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("add", help="append qualified rows to a state sheet")
    a.add_argument("--vertical", required=True)
    a.add_argument("--state", required=True, help="2-letter code, state name, or 'City, ST'")
    a.add_argument("--rows", required=True, help="path to JSON, or '-' for stdin")
    a.set_defaults(func=cmd_add)

    s = sub.add_parser("status", help="show sheets and row counts")
    s.add_argument("--vertical", required=True)
    s.set_defaults(func=cmd_status)

    v = sub.add_parser("versions", help="list restore points")
    v.add_argument("--vertical", required=True)
    v.set_defaults(func=cmd_versions)

    r = sub.add_parser("revert", help="restore the master from a snapshot")
    r.add_argument("--vertical", required=True)
    r.add_argument("--version", help="substring of a restore point name")
    r.set_defaults(func=cmd_revert)

    c = sub.add_parser("consolidate", help="fold city/stray sheets into state sheets")
    c.add_argument("--vertical", required=True)
    c.set_defaults(func=cmd_consolidate)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
